"""Shared utilities for Excel handler."""

from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Styles for Excel export
HEADER_FILL = PatternFill(
    start_color="4472C4", end_color="4472C4", fill_type="solid"
)
HEADER_FONT = Font(bold=True, color="FFFFFF")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def autofit_columns(ws: Worksheet) -> None:
    """Auto-fit column widths based on content."""
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if column_letter is None:
                column_letter = get_column_letter(cell.column)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
