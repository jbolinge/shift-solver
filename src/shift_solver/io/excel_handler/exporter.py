"""Excel exporter for schedule data."""

from pathlib import Path

import openpyxl

from shift_solver.io.excel_handler.utils import (
    BORDER,
    HEADER_FILL,
    HEADER_FONT,
    autofit_columns,
)
from shift_solver.models import Schedule


class ExcelExporter:
    """
    Exports schedule data to professionally formatted Excel files.

    Creates workbooks with multiple sheets:
    - Schedule: Main schedule view
    - Statistics: Summary statistics
    - By Worker: Per-worker assignment view (optional)
    """

    def export_schedule(
        self,
        schedule: Schedule,
        output_path: Path,
        include_worker_view: bool = True,
    ) -> None:
        """
        Export a schedule to an Excel file.

        Args:
            schedule: Schedule to export
            output_path: Path for output file
            include_worker_view: Whether to include per-worker sheet
        """
        # Create parent directories
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()

        # Main schedule sheet
        self._create_schedule_sheet(wb, schedule)

        # Statistics sheet
        self._create_statistics_sheet(wb, schedule)

        # Per-worker view
        if include_worker_view:
            self._create_worker_view_sheet(wb, schedule)

        # Remove default sheet if it's empty
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        wb.save(output_path)

    def _create_schedule_sheet(self, wb: openpyxl.Workbook, schedule: Schedule) -> None:
        """Create the main schedule sheet."""
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = "Schedule"

        # Headers
        headers = ["Period", "Start Date", "End Date", "Worker", "Shift Type", "Date"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER

        # Data
        row = 2
        for period in schedule.periods:
            for worker_id, shifts in period.assignments.items():
                for shift in shifts:
                    ws.cell(row=row, column=1, value=period.period_index + 1)
                    ws.cell(row=row, column=2, value=period.period_start)
                    ws.cell(row=row, column=3, value=period.period_end)
                    ws.cell(row=row, column=4, value=worker_id)
                    ws.cell(row=row, column=5, value=shift.shift_type_id)
                    ws.cell(row=row, column=6, value=shift.date)
                    row += 1

        # Auto-fit columns
        autofit_columns(ws)

    def _create_statistics_sheet(
        self, wb: openpyxl.Workbook, schedule: Schedule
    ) -> None:
        """Create statistics sheet."""
        ws = wb.create_sheet("Statistics")

        # Calculate statistics
        worker_shift_counts: dict[str, dict[str, int]] = {}
        for worker in schedule.workers:
            worker_shift_counts[worker.id] = {}

        for period in schedule.periods:
            for worker_id, shifts in period.assignments.items():
                if worker_id not in worker_shift_counts:
                    worker_shift_counts[worker_id] = {}
                for shift in shifts:
                    shift_type = shift.shift_type_id
                    worker_shift_counts[worker_id][shift_type] = (
                        worker_shift_counts[worker_id].get(shift_type, 0) + 1
                    )

        # Get all shift types
        shift_type_ids = sorted({st.id for st in schedule.shift_types})

        # Headers
        headers = ["Worker"] + shift_type_ids + ["Total"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER

        # Data
        row = 2
        for worker in schedule.workers:
            ws.cell(row=row, column=1, value=worker.id)
            total = 0
            for col, shift_type_id in enumerate(shift_type_ids, start=2):
                count = worker_shift_counts.get(worker.id, {}).get(shift_type_id, 0)
                ws.cell(row=row, column=col, value=count)
                total += count
            ws.cell(row=row, column=len(shift_type_ids) + 2, value=total)
            row += 1

        autofit_columns(ws)

    def _create_worker_view_sheet(
        self, wb: openpyxl.Workbook, schedule: Schedule
    ) -> None:
        """Create per-worker view sheet."""
        ws = wb.create_sheet("By Worker")

        # Headers: Worker, then period columns
        headers = ["Worker"] + [f"P{p.period_index + 1}" for p in schedule.periods]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER

        # Build period assignments lookup
        period_assignments: dict[int, dict[str, list[str]]] = {}
        for period in schedule.periods:
            period_assignments[period.period_index] = {}
            for worker_id, shifts in period.assignments.items():
                period_assignments[period.period_index][worker_id] = [
                    s.shift_type_id for s in shifts
                ]

        # Data rows
        row = 2
        for worker in schedule.workers:
            ws.cell(row=row, column=1, value=f"{worker.name} ({worker.id})")
            for col, period in enumerate(schedule.periods, start=2):
                shift_ids = period_assignments.get(period.period_index, {}).get(
                    worker.id, []
                )
                ws.cell(
                    row=row,
                    column=col,
                    value=", ".join(shift_ids) if shift_ids else "-",
                )
            row += 1

        autofit_columns(ws)
