"""Excel loader for importing scheduling data."""

from pathlib import Path
from typing import Any

import openpyxl

from shift_solver.io.date_utils import parse_date
from shift_solver.io.excel_handler.exceptions import ExcelHandlerError
from shift_solver.models import Availability, SchedulingRequest, Worker
from shift_solver.models.data_models import AVAILABILITY_TYPES, REQUEST_TYPES


class ExcelLoader:
    """
    Loads scheduling data from Excel files.

    Supports multi-sheet workbooks with:
    - Workers sheet
    - Availability sheet
    - Requests sheet

    Can also load from single-sheet files.
    """

    def load_workers(
        self, file_path: Path, sheet_name: str | None = None
    ) -> list[Worker]:
        """
        Load workers from an Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Optional sheet name (defaults to first sheet or "Workers")

        Returns:
            List of Worker objects
        """
        rows = self._read_sheet(file_path, sheet_name or "Workers")
        if not rows:
            return []

        workers = []
        for line_num, row in enumerate(rows, start=2):
            try:
                worker = self._parse_worker_row(row, line_num)
                workers.append(worker)
            except ExcelHandlerError:
                raise  # Already has line number context
            except Exception as e:
                raise ExcelHandlerError(f"Error on row {line_num}: {e}") from e

        return workers

    def load_availability(
        self, file_path: Path, sheet_name: str | None = None
    ) -> list[Availability]:
        """
        Load availability from an Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Optional sheet name (defaults to "Availability")

        Returns:
            List of Availability objects
        """
        rows = self._read_sheet(file_path, sheet_name or "Availability")
        if not rows:
            return []

        availabilities = []
        for line_num, row in enumerate(rows, start=2):
            try:
                avail = self._parse_availability_row(row, line_num)
                availabilities.append(avail)
            except ExcelHandlerError:
                raise  # Already has line number context
            except Exception as e:
                raise ExcelHandlerError(f"Error on row {line_num}: {e}") from e

        return availabilities

    def load_requests(
        self, file_path: Path, sheet_name: str | None = None
    ) -> list[SchedulingRequest]:
        """
        Load scheduling requests from an Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Optional sheet name (defaults to "Requests")

        Returns:
            List of SchedulingRequest objects
        """
        rows = self._read_sheet(file_path, sheet_name or "Requests")
        if not rows:
            return []

        requests = []
        for line_num, row in enumerate(rows, start=2):
            try:
                req = self._parse_request_row(row, line_num)
                requests.append(req)
            except ExcelHandlerError:
                raise  # Already has line number context
            except Exception as e:
                raise ExcelHandlerError(f"Error on row {line_num}: {e}") from e

        return requests

    def load_all(
        self, file_path: Path
    ) -> dict[str, list[Worker] | list[Availability] | list[SchedulingRequest]]:
        """
        Load all data from a multi-sheet workbook.

        Looks for sheets named "Workers", "Availability", "Requests".
        Missing sheets return empty lists.

        Args:
            file_path: Path to Excel file

        Returns:
            Dict with keys "workers", "availability", "requests"
        """
        if not file_path.exists():
            raise ExcelHandlerError(f"File not found: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True)

        result: dict[str, Any] = {
            "workers": [],
            "availability": [],
            "requests": [],
        }

        if "Workers" in wb.sheetnames:
            result["workers"] = self.load_workers(file_path, "Workers")

        if "Availability" in wb.sheetnames:
            result["availability"] = self.load_availability(file_path, "Availability")

        if "Requests" in wb.sheetnames:
            result["requests"] = self.load_requests(file_path, "Requests")

        return result

    def _read_sheet(self, file_path: Path, sheet_name: str) -> list[dict[str, Any]]:
        """Read a sheet and return list of row dicts."""
        if not file_path.exists():
            raise ExcelHandlerError(f"File not found: {file_path}")

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            raise ExcelHandlerError(f"Error reading Excel file: {e}") from e

        # Try to find the sheet (use first sheet if named sheet not found)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        if ws is None:
            return []

        # Get headers from first row
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip().lower() if h else "" for h in rows[0]]

        # Convert remaining rows to dicts
        result = []
        for row in rows[1:]:
            if all(cell is None or cell == "" for cell in row):
                continue  # Skip empty rows
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = value
            result.append(row_dict)

        return result

    def _parse_worker_row(self, row: dict[str, Any], line_num: int) -> Worker:
        """Parse a worker row from Excel."""
        worker_id = str(row.get("id") or "").strip()
        name = str(row.get("name") or "").strip()

        if not worker_id:
            raise ExcelHandlerError(f"empty 'id' on line {line_num}")
        if not name:
            raise ExcelHandlerError(f"empty 'name' on line {line_num}")

        worker_type = str(row.get("worker_type", "") or "").strip() or None

        restricted_str = str(row.get("restricted_shifts", "") or "").strip()
        restricted_shifts = frozenset(
            s.strip() for s in restricted_str.split(",") if s.strip()
        )

        preferred_str = str(row.get("preferred_shifts", "") or "").strip()
        preferred_shifts = frozenset(
            s.strip() for s in preferred_str.split(",") if s.strip()
        )

        return Worker(
            id=worker_id,
            name=name,
            worker_type=worker_type,
            restricted_shifts=restricted_shifts,
            preferred_shifts=preferred_shifts,
        )

    def _parse_availability_row(
        self, row: dict[str, Any], line_num: int
    ) -> Availability:
        """Parse an availability row from Excel."""
        worker_id = str(row.get("worker_id") or "").strip()
        if not worker_id:
            raise ExcelHandlerError(f"empty 'worker_id' on line {line_num}")

        start_date = parse_date(
            row.get("start_date"), "start_date", line_num, ExcelHandlerError
        )
        end_date = parse_date(
            row.get("end_date"), "end_date", line_num, ExcelHandlerError
        )

        availability_type = str(row.get("availability_type") or "").strip()
        if availability_type not in AVAILABILITY_TYPES:
            raise ExcelHandlerError(
                f"Invalid availability_type '{availability_type}' on line {line_num}"
            )

        shift_type_id = str(row.get("shift_type_id", "") or "").strip() or None

        return Availability(
            worker_id=worker_id,
            start_date=start_date,
            end_date=end_date,
            availability_type=availability_type,  # type: ignore
            shift_type_id=shift_type_id,
        )

    def _parse_priority(self, priority_val: Any, line_num: int) -> int:
        """Parse and validate priority value.

        Args:
            priority_val: Priority value from Excel cell (can be int, float, str, or None)
            line_num: Line number for error messages

        Returns:
            Validated integer priority

        Raises:
            ExcelHandlerError: If priority is invalid
        """
        # Handle empty/None values
        if priority_val is None or priority_val == "":
            return 1  # Default priority

        # Handle numeric values (int or float from Excel)
        if isinstance(priority_val, (int, float)):
            # Reject floats that aren't whole numbers
            if isinstance(priority_val, float) and priority_val != int(priority_val):
                raise ExcelHandlerError(
                    f"Invalid priority value '{priority_val}' on line {line_num}. "
                    f"Must be a positive integer."
                )
            priority = int(priority_val)
        else:
            # Handle string values
            priority_str = str(priority_val).strip()
            if not priority_str:
                return 1  # Default priority

            try:
                priority = int(priority_str)
            except ValueError as e:
                raise ExcelHandlerError(
                    f"Invalid priority value '{priority_str}' on line {line_num}. "
                    f"Must be a positive integer."
                ) from e

        if priority <= 0:
            raise ExcelHandlerError(
                f"priority must be positive, got '{priority}' on line {line_num}"
            )

        return priority

    def _parse_request_row(
        self, row: dict[str, Any], line_num: int
    ) -> SchedulingRequest:
        """Parse a request row from Excel."""
        worker_id = str(row.get("worker_id") or "").strip()
        if not worker_id:
            raise ExcelHandlerError(f"empty 'worker_id' on line {line_num}")

        start_date = parse_date(
            row.get("start_date"), "start_date", line_num, ExcelHandlerError
        )
        end_date = parse_date(
            row.get("end_date"), "end_date", line_num, ExcelHandlerError
        )

        request_type = str(row.get("request_type") or "").strip()
        if request_type not in REQUEST_TYPES:
            raise ExcelHandlerError(
                f"Invalid request_type '{request_type}' on line {line_num}"
            )

        shift_type_id = str(row.get("shift_type_id") or "").strip()
        if not shift_type_id:
            raise ExcelHandlerError(f"empty 'shift_type_id' on line {line_num}")

        priority = self._parse_priority(row.get("priority"), line_num)

        return SchedulingRequest(
            worker_id=worker_id,
            start_date=start_date,
            end_date=end_date,
            request_type=request_type,  # type: ignore
            shift_type_id=shift_type_id,
            priority=priority,
        )
