"""Export functions for sample data."""

import csv
from datetime import date
from pathlib import Path
from typing import TYPE_CHECKING, Protocol

import openpyxl

from shift_solver.models import Availability, SchedulingRequest, ShiftType, Worker

if TYPE_CHECKING:
    pass


class _HasGenerateMethods(Protocol):
    """Protocol for classes with generate methods."""

    def generate_workers(self, num_workers: int) -> list[Worker]: ...
    def generate_shift_types(self) -> list[ShiftType]: ...
    def generate_availability(
        self, workers: list[Worker], start_date: date, end_date: date
    ) -> list[Availability]: ...
    def generate_requests(
        self,
        workers: list[Worker],
        shift_types: list[ShiftType],
        start_date: date,
        end_date: date,
    ) -> list[SchedulingRequest]: ...


def export_to_csv(
    output_dir: Path,
    workers: list[Worker],
    shift_types: list[ShiftType],
    availability: list[Availability],
    requests: list[SchedulingRequest],
) -> None:
    """
    Export sample data to CSV files.

    Creates:
    - workers.csv
    - shift_types.csv
    - availability.csv
    - requests.csv

    Args:
        output_dir: Directory to write files to
        workers: List of workers
        shift_types: List of shift types
        availability: List of availability records
        requests: List of scheduling requests
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    # Write workers.csv
    with open(output_dir / "workers.csv", "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["id", "name", "worker_type", "restricted_shifts"])
        for w in workers:
            writer.writerow(
                [
                    w.id,
                    w.name,
                    w.worker_type or "",
                    ",".join(w.restricted_shifts),
                ]
            )

    # Write shift_types.csv
    with open(output_dir / "shift_types.csv", "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "id",
                "name",
                "category",
                "start_time",
                "end_time",
                "duration_hours",
                "workers_required",
                "is_undesirable",
            ]
        )
        for st in shift_types:
            writer.writerow(
                [
                    st.id,
                    st.name,
                    st.category,
                    st.start_time.strftime("%H:%M"),
                    st.end_time.strftime("%H:%M"),
                    st.duration_hours,
                    st.workers_required,
                    str(st.is_undesirable).lower(),
                ]
            )

    # Write availability.csv
    with open(output_dir / "availability.csv", "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "worker_id",
                "start_date",
                "end_date",
                "availability_type",
                "shift_type_id",
            ]
        )
        for a in availability:
            writer.writerow(
                [
                    a.worker_id,
                    a.start_date.isoformat(),
                    a.end_date.isoformat(),
                    a.availability_type,
                    a.shift_type_id or "",
                ]
            )

    # Write requests.csv
    with open(output_dir / "requests.csv", "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "worker_id",
                "start_date",
                "end_date",
                "request_type",
                "shift_type_id",
                "priority",
            ]
        )
        for r in requests:
            writer.writerow(
                [
                    r.worker_id,
                    r.start_date.isoformat(),
                    r.end_date.isoformat(),
                    r.request_type,
                    r.shift_type_id,
                    r.priority,
                ]
            )


def export_to_excel(
    output_file: Path,
    workers: list[Worker],
    shift_types: list[ShiftType],
    availability: list[Availability],
    requests: list[SchedulingRequest],
) -> None:
    """
    Export sample data to an Excel file.

    Creates a workbook with sheets:
    - Workers
    - ShiftTypes
    - Availability
    - Requests

    Args:
        output_file: Path for output Excel file
        workers: List of workers
        shift_types: List of shift types
        availability: List of availability records
        requests: List of scheduling requests
    """
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # Workers sheet
    ws_workers = wb.active
    ws_workers.title = "Workers"
    ws_workers.append(["id", "name", "worker_type", "restricted_shifts"])
    for w in workers:
        ws_workers.append(
            [
                w.id,
                w.name,
                w.worker_type or "",
                ",".join(w.restricted_shifts),
            ]
        )

    # ShiftTypes sheet
    ws_shifts = wb.create_sheet("ShiftTypes")
    ws_shifts.append(
        [
            "id",
            "name",
            "category",
            "start_time",
            "end_time",
            "duration_hours",
            "workers_required",
            "is_undesirable",
        ]
    )
    for st in shift_types:
        ws_shifts.append(
            [
                st.id,
                st.name,
                st.category,
                st.start_time.strftime("%H:%M"),
                st.end_time.strftime("%H:%M"),
                st.duration_hours,
                st.workers_required,
                st.is_undesirable,
            ]
        )

    # Availability sheet
    ws_avail = wb.create_sheet("Availability")
    ws_avail.append(
        [
            "worker_id",
            "start_date",
            "end_date",
            "availability_type",
            "shift_type_id",
        ]
    )
    for a in availability:
        ws_avail.append(
            [
                a.worker_id,
                a.start_date,
                a.end_date,
                a.availability_type,
                a.shift_type_id or "",
            ]
        )

    # Requests sheet
    ws_req = wb.create_sheet("Requests")
    ws_req.append(
        [
            "worker_id",
            "start_date",
            "end_date",
            "request_type",
            "shift_type_id",
            "priority",
        ]
    )
    for r in requests:
        ws_req.append(
            [
                r.worker_id,
                r.start_date,
                r.end_date,
                r.request_type,
                r.shift_type_id,
                r.priority,
            ]
        )

    wb.save(output_file)


class SampleExporterMixin(_HasGenerateMethods):
    """Mixin that adds export methods to SampleGenerator."""

    def generate_to_csv(
        self: _HasGenerateMethods,
        output_dir: Path,
        num_workers: int,
        start_date: date,
        end_date: date,
    ) -> None:
        """
        Generate sample data and write to CSV files.

        Creates:
        - workers.csv
        - shift_types.csv
        - availability.csv
        - requests.csv

        Args:
            output_dir: Directory to write files to
            num_workers: Number of workers to generate
            start_date: Schedule start date
            end_date: Schedule end date
        """
        # Generate data
        workers = self.generate_workers(num_workers)
        shift_types = self.generate_shift_types()
        availability = self.generate_availability(workers, start_date, end_date)
        requests = self.generate_requests(workers, shift_types, start_date, end_date)

        export_to_csv(output_dir, workers, shift_types, availability, requests)

    def generate_to_excel(
        self: _HasGenerateMethods,
        output_file: Path,
        num_workers: int,
        start_date: date,
        end_date: date,
    ) -> None:
        """
        Generate sample data and write to Excel file.

        Creates a workbook with sheets:
        - Workers
        - ShiftTypes
        - Availability
        - Requests

        Args:
            output_file: Path for output Excel file
            num_workers: Number of workers to generate
            start_date: Schedule start date
            end_date: Schedule end date
        """
        # Generate data
        workers = self.generate_workers(num_workers)
        shift_types = self.generate_shift_types()
        availability = self.generate_availability(workers, start_date, end_date)
        requests = self.generate_requests(workers, shift_types, start_date, end_date)

        export_to_excel(output_file, workers, shift_types, availability, requests)
