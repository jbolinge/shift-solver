"""Sample data generator for different industries."""

import csv
import random
from dataclasses import dataclass
from datetime import date, time, timedelta
from pathlib import Path
from typing import Any

import openpyxl

from shift_solver.models import Availability, SchedulingRequest, ShiftType, Worker


@dataclass
class IndustryPreset:
    """Configuration preset for an industry."""

    name: str
    shift_types: list[dict[str, Any]]
    worker_types: list[str]
    restriction_probability: float = 0.1
    vacation_probability: float = 0.15
    request_probability: float = 0.2

    @classmethod
    def get(cls, industry: str) -> "IndustryPreset":
        """Get preset for an industry."""
        presets = {
            "retail": cls._retail_preset(),
            "healthcare": cls._healthcare_preset(),
            "warehouse": cls._warehouse_preset(),
        }
        if industry not in presets:
            raise ValueError(
                f"Unknown industry '{industry}'. Available: {list(presets.keys())}"
            )
        return presets[industry]

    @classmethod
    def _retail_preset(cls) -> "IndustryPreset":
        """Retail industry preset."""
        return cls(
            name="retail",
            shift_types=[
                {
                    "id": "morning",
                    "name": "Morning Shift",
                    "category": "day",
                    "start_time": time(6, 0),
                    "end_time": time(14, 0),
                    "duration_hours": 8.0,
                    "workers_required": 3,
                    "is_undesirable": False,
                },
                {
                    "id": "afternoon",
                    "name": "Afternoon Shift",
                    "category": "day",
                    "start_time": time(14, 0),
                    "end_time": time(22, 0),
                    "duration_hours": 8.0,
                    "workers_required": 4,
                    "is_undesirable": False,
                },
                {
                    "id": "night",
                    "name": "Night Shift",
                    "category": "night",
                    "start_time": time(22, 0),
                    "end_time": time(6, 0),
                    "duration_hours": 8.0,
                    "workers_required": 2,
                    "is_undesirable": True,
                },
                {
                    "id": "weekend",
                    "name": "Weekend Shift",
                    "category": "weekend",
                    "start_time": time(10, 0),
                    "end_time": time(18, 0),
                    "duration_hours": 8.0,
                    "workers_required": 5,
                    "is_undesirable": True,
                },
            ],
            worker_types=["full_time", "part_time", "seasonal"],
            restriction_probability=0.15,
            vacation_probability=0.1,
            request_probability=0.15,
        )

    @classmethod
    def _healthcare_preset(cls) -> "IndustryPreset":
        """Healthcare industry preset."""
        return cls(
            name="healthcare",
            shift_types=[
                {
                    "id": "day",
                    "name": "Day Shift",
                    "category": "day",
                    "start_time": time(7, 0),
                    "end_time": time(19, 0),
                    "duration_hours": 12.0,
                    "workers_required": 4,
                    "is_undesirable": False,
                },
                {
                    "id": "night",
                    "name": "Night Shift",
                    "category": "night",
                    "start_time": time(19, 0),
                    "end_time": time(7, 0),
                    "duration_hours": 12.0,
                    "workers_required": 3,
                    "is_undesirable": True,
                },
                {
                    "id": "on_call",
                    "name": "On-Call",
                    "category": "on_call",
                    "start_time": time(0, 0),
                    "end_time": time(23, 59),
                    "duration_hours": 24.0,
                    "workers_required": 1,
                    "is_undesirable": True,
                },
            ],
            worker_types=["physician", "nurse", "resident"],
            restriction_probability=0.2,
            vacation_probability=0.12,
            request_probability=0.25,
        )

    @classmethod
    def _warehouse_preset(cls) -> "IndustryPreset":
        """Warehouse industry preset."""
        return cls(
            name="warehouse",
            shift_types=[
                {
                    "id": "first",
                    "name": "First Shift",
                    "category": "day",
                    "start_time": time(6, 0),
                    "end_time": time(14, 0),
                    "duration_hours": 8.0,
                    "workers_required": 8,
                    "is_undesirable": False,
                },
                {
                    "id": "second",
                    "name": "Second Shift",
                    "category": "evening",
                    "start_time": time(14, 0),
                    "end_time": time(22, 0),
                    "duration_hours": 8.0,
                    "workers_required": 6,
                    "is_undesirable": False,
                },
                {
                    "id": "third",
                    "name": "Third Shift",
                    "category": "night",
                    "start_time": time(22, 0),
                    "end_time": time(6, 0),
                    "duration_hours": 8.0,
                    "workers_required": 4,
                    "is_undesirable": True,
                },
            ],
            worker_types=["forklift_operator", "picker", "supervisor"],
            restriction_probability=0.1,
            vacation_probability=0.08,
            request_probability=0.1,
        )


# Common first and last names for generating worker names
FIRST_NAMES = [
    "James",
    "Mary",
    "John",
    "Patricia",
    "Robert",
    "Jennifer",
    "Michael",
    "Linda",
    "William",
    "Elizabeth",
    "David",
    "Barbara",
    "Richard",
    "Susan",
    "Joseph",
    "Jessica",
    "Thomas",
    "Sarah",
    "Charles",
    "Karen",
    "Christopher",
    "Lisa",
    "Daniel",
    "Nancy",
    "Matthew",
    "Betty",
    "Anthony",
    "Margaret",
    "Mark",
    "Sandra",
    "Donald",
    "Ashley",
    "Steven",
    "Kimberly",
    "Paul",
    "Emily",
    "Andrew",
    "Donna",
    "Joshua",
    "Michelle",
]

LAST_NAMES = [
    "Smith",
    "Johnson",
    "Williams",
    "Brown",
    "Jones",
    "Garcia",
    "Miller",
    "Davis",
    "Rodriguez",
    "Martinez",
    "Hernandez",
    "Lopez",
    "Gonzalez",
    "Wilson",
    "Anderson",
    "Thomas",
    "Taylor",
    "Moore",
    "Jackson",
    "Martin",
    "Lee",
    "Perez",
    "Thompson",
    "White",
    "Harris",
    "Sanchez",
    "Clark",
    "Ramirez",
    "Lewis",
    "Robinson",
    "Walker",
]


class SampleGenerator:
    """
    Generates sample scheduling data for different industries.

    Supports generating:
    - Workers with realistic names and types
    - Shift types based on industry presets
    - Availability records (vacations, time-off)
    - Scheduling requests

    Usage:
        gen = SampleGenerator(industry="retail", seed=42)
        workers = gen.generate_workers(num_workers=20)
        shift_types = gen.generate_shift_types()
        gen.generate_to_csv(output_dir, num_workers=20, start_date, end_date)
    """

    def __init__(self, industry: str = "retail", seed: int | None = None) -> None:
        """
        Initialize sample generator.

        Args:
            industry: Industry preset to use (retail, healthcare, warehouse)
            seed: Random seed for reproducibility
        """
        self.preset = IndustryPreset.get(industry)
        self.rng = random.Random(seed)

    def generate_workers(self, num_workers: int) -> list[Worker]:
        """
        Generate sample workers.

        Args:
            num_workers: Number of workers to generate

        Returns:
            List of Worker objects
        """
        workers = []
        used_names: set[str] = set()

        for i in range(num_workers):
            worker_id = f"W{i + 1:03d}"

            # Generate unique name
            name = self._generate_unique_name(used_names)
            used_names.add(name)

            # Random worker type
            worker_type = self.rng.choice(self.preset.worker_types)

            # Random restrictions
            restricted_shifts: frozenset[str] = frozenset()
            if self.rng.random() < self.preset.restriction_probability:
                # Pick 1-2 shifts to restrict
                shift_ids = [st["id"] for st in self.preset.shift_types]
                num_restrictions = self.rng.randint(1, min(2, len(shift_ids)))
                restricted_shifts = frozenset(
                    self.rng.sample(shift_ids, num_restrictions)
                )

            workers.append(
                Worker(
                    id=worker_id,
                    name=name,
                    worker_type=worker_type,
                    restricted_shifts=restricted_shifts,
                )
            )

        return workers

    def generate_shift_types(self) -> list[ShiftType]:
        """
        Generate shift types from industry preset.

        Returns:
            List of ShiftType objects
        """
        return [
            ShiftType(
                id=st["id"],
                name=st["name"],
                category=st["category"],
                start_time=st["start_time"],
                end_time=st["end_time"],
                duration_hours=st["duration_hours"],
                workers_required=st["workers_required"],
                is_undesirable=st.get("is_undesirable", False),
            )
            for st in self.preset.shift_types
        ]

    def generate_availability(
        self,
        workers: list[Worker],
        start_date: date,
        end_date: date,
    ) -> list[Availability]:
        """
        Generate availability records (vacations, time-off).

        Args:
            workers: List of workers
            start_date: Schedule start date
            end_date: Schedule end date

        Returns:
            List of Availability objects (unavailable periods)
        """
        availabilities = []
        total_days = (end_date - start_date).days + 1

        for worker in workers:
            # Decide if worker has vacation
            if self.rng.random() < self.preset.vacation_probability:
                # Generate 1-2 vacation periods
                num_vacations = self.rng.randint(1, 2)

                for _ in range(num_vacations):
                    # Random vacation length (3-10 days)
                    vacation_length = self.rng.randint(3, 10)

                    # Random start within the period
                    max_start_offset = max(0, total_days - vacation_length - 1)
                    if max_start_offset <= 0:
                        continue

                    start_offset = self.rng.randint(0, max_start_offset)
                    vac_start = start_date + timedelta(days=start_offset)
                    vac_end = vac_start + timedelta(days=vacation_length - 1)

                    availabilities.append(
                        Availability(
                            worker_id=worker.id,
                            start_date=vac_start,
                            end_date=vac_end,
                            availability_type="unavailable",
                        )
                    )

        return availabilities

    def generate_requests(
        self,
        workers: list[Worker],
        shift_types: list[ShiftType],
        start_date: date,
        end_date: date,
    ) -> list[SchedulingRequest]:
        """
        Generate scheduling requests.

        Args:
            workers: List of workers
            shift_types: List of shift types
            start_date: Schedule start date
            end_date: Schedule end date

        Returns:
            List of SchedulingRequest objects
        """
        requests = []
        total_days = (end_date - start_date).days + 1

        for worker in workers:
            # Decide if worker has requests
            if self.rng.random() < self.preset.request_probability:
                # Generate 1-3 requests
                num_requests = self.rng.randint(1, 3)

                for _ in range(num_requests):
                    # Random request type
                    request_type = self.rng.choice(["positive", "negative"])

                    # Random shift type
                    shift_type = self.rng.choice(shift_types)

                    # Random date
                    day_offset = self.rng.randint(0, total_days - 1)
                    req_date = start_date + timedelta(days=day_offset)

                    # Random priority (1-3)
                    priority = self.rng.randint(1, 3)

                    requests.append(
                        SchedulingRequest(
                            worker_id=worker.id,
                            start_date=req_date,
                            end_date=req_date,
                            request_type=request_type,  # type: ignore
                            shift_type_id=shift_type.id,
                            priority=priority,
                        )
                    )

        return requests

    def generate_to_csv(
        self,
        output_dir: Path,
        num_workers: int,
        start_date: date,
        end_date: date,
    ) -> None:
        """
        Generate sample data and write to CSV files.

        Creates:
        - workers.csv
        - shift_types.csv
        - availability.csv
        - requests.csv

        Args:
            output_dir: Directory to write files to
            num_workers: Number of workers to generate
            start_date: Schedule start date
            end_date: Schedule end date
        """
        output_dir.mkdir(parents=True, exist_ok=True)

        # Generate data
        workers = self.generate_workers(num_workers)
        shift_types = self.generate_shift_types()
        availability = self.generate_availability(workers, start_date, end_date)
        requests = self.generate_requests(workers, shift_types, start_date, end_date)

        # Write workers.csv
        with open(output_dir / "workers.csv", "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["id", "name", "worker_type", "restricted_shifts"])
            for w in workers:
                writer.writerow(
                    [
                        w.id,
                        w.name,
                        w.worker_type or "",
                        ",".join(w.restricted_shifts),
                    ]
                )

        # Write shift_types.csv
        with open(output_dir / "shift_types.csv", "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "id",
                    "name",
                    "category",
                    "start_time",
                    "end_time",
                    "duration_hours",
                    "workers_required",
                    "is_undesirable",
                ]
            )
            for st in shift_types:
                writer.writerow(
                    [
                        st.id,
                        st.name,
                        st.category,
                        st.start_time.strftime("%H:%M"),
                        st.end_time.strftime("%H:%M"),
                        st.duration_hours,
                        st.workers_required,
                        str(st.is_undesirable).lower(),
                    ]
                )

        # Write availability.csv
        with open(output_dir / "availability.csv", "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "worker_id",
                    "start_date",
                    "end_date",
                    "availability_type",
                    "shift_type_id",
                ]
            )
            for a in availability:
                writer.writerow(
                    [
                        a.worker_id,
                        a.start_date.isoformat(),
                        a.end_date.isoformat(),
                        a.availability_type,
                        a.shift_type_id or "",
                    ]
                )

        # Write requests.csv
        with open(output_dir / "requests.csv", "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "worker_id",
                    "start_date",
                    "end_date",
                    "request_type",
                    "shift_type_id",
                    "priority",
                ]
            )
            for r in requests:
                writer.writerow(
                    [
                        r.worker_id,
                        r.start_date.isoformat(),
                        r.end_date.isoformat(),
                        r.request_type,
                        r.shift_type_id,
                        r.priority,
                    ]
                )

    def generate_to_excel(
        self,
        output_file: Path,
        num_workers: int,
        start_date: date,
        end_date: date,
    ) -> None:
        """
        Generate sample data and write to Excel file.

        Creates a workbook with sheets:
        - Workers
        - ShiftTypes
        - Availability
        - Requests

        Args:
            output_file: Path for output Excel file
            num_workers: Number of workers to generate
            start_date: Schedule start date
            end_date: Schedule end date
        """
        output_file.parent.mkdir(parents=True, exist_ok=True)

        # Generate data
        workers = self.generate_workers(num_workers)
        shift_types = self.generate_shift_types()
        availability = self.generate_availability(workers, start_date, end_date)
        requests = self.generate_requests(workers, shift_types, start_date, end_date)

        wb = openpyxl.Workbook()

        # Workers sheet
        ws_workers = wb.active
        ws_workers.title = "Workers"
        ws_workers.append(["id", "name", "worker_type", "restricted_shifts"])
        for w in workers:
            ws_workers.append(
                [
                    w.id,
                    w.name,
                    w.worker_type or "",
                    ",".join(w.restricted_shifts),
                ]
            )

        # ShiftTypes sheet
        ws_shifts = wb.create_sheet("ShiftTypes")
        ws_shifts.append(
            [
                "id",
                "name",
                "category",
                "start_time",
                "end_time",
                "duration_hours",
                "workers_required",
                "is_undesirable",
            ]
        )
        for st in shift_types:
            ws_shifts.append(
                [
                    st.id,
                    st.name,
                    st.category,
                    st.start_time.strftime("%H:%M"),
                    st.end_time.strftime("%H:%M"),
                    st.duration_hours,
                    st.workers_required,
                    st.is_undesirable,
                ]
            )

        # Availability sheet
        ws_avail = wb.create_sheet("Availability")
        ws_avail.append(
            [
                "worker_id",
                "start_date",
                "end_date",
                "availability_type",
                "shift_type_id",
            ]
        )
        for a in availability:
            ws_avail.append(
                [
                    a.worker_id,
                    a.start_date,
                    a.end_date,
                    a.availability_type,
                    a.shift_type_id or "",
                ]
            )

        # Requests sheet
        ws_req = wb.create_sheet("Requests")
        ws_req.append(
            [
                "worker_id",
                "start_date",
                "end_date",
                "request_type",
                "shift_type_id",
                "priority",
            ]
        )
        for r in requests:
            ws_req.append(
                [
                    r.worker_id,
                    r.start_date,
                    r.end_date,
                    r.request_type,
                    r.shift_type_id,
                    r.priority,
                ]
            )

        wb.save(output_file)

    def _generate_unique_name(self, used_names: set[str]) -> str:
        """Generate a unique full name."""
        for _ in range(100):  # Max attempts
            first = self.rng.choice(FIRST_NAMES)
            last = self.rng.choice(LAST_NAMES)
            name = f"{first} {last}"
            if name not in used_names:
                return name

        # Fallback with number suffix
        first = self.rng.choice(FIRST_NAMES)
        last = self.rng.choice(LAST_NAMES)
        return f"{first} {last} {len(used_names)}"
