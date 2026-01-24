"""Excel import/export handler for scheduling data."""

from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from shift_solver.models import (
    Availability,
    Schedule,
    SchedulingRequest,
    Worker,
)
from shift_solver.models.data_models import AVAILABILITY_TYPES, REQUEST_TYPES


class ExcelHandlerError(Exception):
    """Error during Excel import/export."""

    pass


class ExcelLoader:
    """
    Loads scheduling data from Excel files.

    Supports multi-sheet workbooks with:
    - Workers sheet
    - Availability sheet
    - Requests sheet

    Can also load from single-sheet files.
    """

    def load_workers(
        self, file_path: Path, sheet_name: str | None = None
    ) -> list[Worker]:
        """
        Load workers from an Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Optional sheet name (defaults to first sheet or "Workers")

        Returns:
            List of Worker objects
        """
        rows = self._read_sheet(file_path, sheet_name or "Workers")
        if not rows:
            return []

        workers = []
        for line_num, row in enumerate(rows, start=2):
            worker = self._parse_worker_row(row, line_num)
            workers.append(worker)

        return workers

    def load_availability(
        self, file_path: Path, sheet_name: str | None = None
    ) -> list[Availability]:
        """
        Load availability from an Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Optional sheet name (defaults to "Availability")

        Returns:
            List of Availability objects
        """
        rows = self._read_sheet(file_path, sheet_name or "Availability")
        if not rows:
            return []

        availabilities = []
        for line_num, row in enumerate(rows, start=2):
            avail = self._parse_availability_row(row, line_num)
            availabilities.append(avail)

        return availabilities

    def load_requests(
        self, file_path: Path, sheet_name: str | None = None
    ) -> list[SchedulingRequest]:
        """
        Load scheduling requests from an Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Optional sheet name (defaults to "Requests")

        Returns:
            List of SchedulingRequest objects
        """
        rows = self._read_sheet(file_path, sheet_name or "Requests")
        if not rows:
            return []

        requests = []
        for line_num, row in enumerate(rows, start=2):
            req = self._parse_request_row(row, line_num)
            requests.append(req)

        return requests

    def load_all(
        self, file_path: Path
    ) -> dict[str, list[Worker] | list[Availability] | list[SchedulingRequest]]:
        """
        Load all data from a multi-sheet workbook.

        Looks for sheets named "Workers", "Availability", "Requests".
        Missing sheets return empty lists.

        Args:
            file_path: Path to Excel file

        Returns:
            Dict with keys "workers", "availability", "requests"
        """
        if not file_path.exists():
            raise ExcelHandlerError(f"File not found: {file_path}")

        wb = openpyxl.load_workbook(file_path, data_only=True)

        result: dict[str, Any] = {
            "workers": [],
            "availability": [],
            "requests": [],
        }

        if "Workers" in wb.sheetnames:
            result["workers"] = self.load_workers(file_path, "Workers")

        if "Availability" in wb.sheetnames:
            result["availability"] = self.load_availability(file_path, "Availability")

        if "Requests" in wb.sheetnames:
            result["requests"] = self.load_requests(file_path, "Requests")

        return result

    def _read_sheet(self, file_path: Path, sheet_name: str) -> list[dict[str, Any]]:
        """Read a sheet and return list of row dicts."""
        if not file_path.exists():
            raise ExcelHandlerError(f"File not found: {file_path}")

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            raise ExcelHandlerError(f"Error reading Excel file: {e}") from e

        # Try to find the sheet (use first sheet if named sheet not found)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        if ws is None:
            return []

        # Get headers from first row
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip().lower() if h else "" for h in rows[0]]

        # Convert remaining rows to dicts
        result = []
        for row in rows[1:]:
            if all(cell is None or cell == "" for cell in row):
                continue  # Skip empty rows
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = value
            result.append(row_dict)

        return result

    def _parse_worker_row(self, row: dict[str, Any], line_num: int) -> Worker:
        """Parse a worker row from Excel."""
        worker_id = str(row.get("id", "")).strip()
        name = str(row.get("name", "")).strip()

        if not worker_id:
            raise ExcelHandlerError(f"empty 'id' on line {line_num}")
        if not name:
            raise ExcelHandlerError(f"empty 'name' on line {line_num}")

        worker_type = str(row.get("worker_type", "") or "").strip() or None

        restricted_str = str(row.get("restricted_shifts", "") or "").strip()
        restricted_shifts = frozenset(
            s.strip() for s in restricted_str.split(",") if s.strip()
        )

        preferred_str = str(row.get("preferred_shifts", "") or "").strip()
        preferred_shifts = frozenset(
            s.strip() for s in preferred_str.split(",") if s.strip()
        )

        return Worker(
            id=worker_id,
            name=name,
            worker_type=worker_type,
            restricted_shifts=restricted_shifts,
            preferred_shifts=preferred_shifts,
        )

    def _parse_availability_row(
        self, row: dict[str, Any], line_num: int
    ) -> Availability:
        """Parse an availability row from Excel."""
        worker_id = str(row.get("worker_id", "")).strip()
        if not worker_id:
            raise ExcelHandlerError(f"empty 'worker_id' on line {line_num}")

        start_date = self._parse_date(row.get("start_date"), "start_date", line_num)
        end_date = self._parse_date(row.get("end_date"), "end_date", line_num)

        availability_type = str(row.get("availability_type", "")).strip()
        if availability_type not in AVAILABILITY_TYPES:
            raise ExcelHandlerError(
                f"Invalid availability_type '{availability_type}' on line {line_num}"
            )

        shift_type_id = str(row.get("shift_type_id", "") or "").strip() or None

        return Availability(
            worker_id=worker_id,
            start_date=start_date,
            end_date=end_date,
            availability_type=availability_type,  # type: ignore
            shift_type_id=shift_type_id,
        )

    def _parse_request_row(
        self, row: dict[str, Any], line_num: int
    ) -> SchedulingRequest:
        """Parse a request row from Excel."""
        worker_id = str(row.get("worker_id", "")).strip()
        if not worker_id:
            raise ExcelHandlerError(f"empty 'worker_id' on line {line_num}")

        start_date = self._parse_date(row.get("start_date"), "start_date", line_num)
        end_date = self._parse_date(row.get("end_date"), "end_date", line_num)

        request_type = str(row.get("request_type", "")).strip()
        if request_type not in REQUEST_TYPES:
            raise ExcelHandlerError(
                f"Invalid request_type '{request_type}' on line {line_num}"
            )

        shift_type_id = str(row.get("shift_type_id", "")).strip()
        if not shift_type_id:
            raise ExcelHandlerError(f"empty 'shift_type_id' on line {line_num}")

        priority_val = row.get("priority")
        priority = int(priority_val) if priority_val else 1

        return SchedulingRequest(
            worker_id=worker_id,
            start_date=start_date,
            end_date=end_date,
            request_type=request_type,  # type: ignore
            shift_type_id=shift_type_id,
            priority=priority,
        )

    def _parse_date(self, value: Any, field_name: str, line_num: int) -> date:
        """Parse a date from Excel cell value."""
        if value is None:
            raise ExcelHandlerError(f"empty '{field_name}' on line {line_num}")

        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value

        # Try to parse string
        date_str = str(value).strip()
        for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"]:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue

        raise ExcelHandlerError(
            f"Invalid date '{value}' for '{field_name}' on line {line_num}"
        )


class ExcelExporter:
    """
    Exports schedule data to professionally formatted Excel files.

    Creates workbooks with multiple sheets:
    - Schedule: Main schedule view
    - Statistics: Summary statistics
    - By Worker: Per-worker assignment view (optional)
    """

    # Styles
    HEADER_FILL = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def export_schedule(
        self,
        schedule: Schedule,
        output_path: Path,
        include_worker_view: bool = True,
    ) -> None:
        """
        Export a schedule to an Excel file.

        Args:
            schedule: Schedule to export
            output_path: Path for output file
            include_worker_view: Whether to include per-worker sheet
        """
        # Create parent directories
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()

        # Main schedule sheet
        self._create_schedule_sheet(wb, schedule)

        # Statistics sheet
        self._create_statistics_sheet(wb, schedule)

        # Per-worker view
        if include_worker_view:
            self._create_worker_view_sheet(wb, schedule)

        # Remove default sheet if it's empty
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        wb.save(output_path)

    def _create_schedule_sheet(self, wb: openpyxl.Workbook, schedule: Schedule) -> None:
        """Create the main schedule sheet."""
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = "Schedule"

        # Headers
        headers = ["Period", "Start Date", "End Date", "Worker", "Shift Type", "Date"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER

        # Data
        row = 2
        for period in schedule.periods:
            for worker_id, shifts in period.assignments.items():
                for shift in shifts:
                    ws.cell(row=row, column=1, value=period.period_index + 1)
                    ws.cell(row=row, column=2, value=period.period_start)
                    ws.cell(row=row, column=3, value=period.period_end)
                    ws.cell(row=row, column=4, value=worker_id)
                    ws.cell(row=row, column=5, value=shift.shift_type_id)
                    ws.cell(row=row, column=6, value=shift.date)
                    row += 1

        # Auto-fit columns
        self._autofit_columns(ws)

    def _create_statistics_sheet(
        self, wb: openpyxl.Workbook, schedule: Schedule
    ) -> None:
        """Create statistics sheet."""
        ws = wb.create_sheet("Statistics")

        # Calculate statistics
        worker_shift_counts: dict[str, dict[str, int]] = {}
        for worker in schedule.workers:
            worker_shift_counts[worker.id] = {}

        for period in schedule.periods:
            for worker_id, shifts in period.assignments.items():
                if worker_id not in worker_shift_counts:
                    worker_shift_counts[worker_id] = {}
                for shift in shifts:
                    shift_type = shift.shift_type_id
                    worker_shift_counts[worker_id][shift_type] = (
                        worker_shift_counts[worker_id].get(shift_type, 0) + 1
                    )

        # Get all shift types
        shift_type_ids = sorted({st.id for st in schedule.shift_types})

        # Headers
        headers = ["Worker"] + shift_type_ids + ["Total"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER

        # Data
        row = 2
        for worker in schedule.workers:
            ws.cell(row=row, column=1, value=worker.id)
            total = 0
            for col, shift_type_id in enumerate(shift_type_ids, start=2):
                count = worker_shift_counts.get(worker.id, {}).get(shift_type_id, 0)
                ws.cell(row=row, column=col, value=count)
                total += count
            ws.cell(row=row, column=len(shift_type_ids) + 2, value=total)
            row += 1

        self._autofit_columns(ws)

    def _create_worker_view_sheet(
        self, wb: openpyxl.Workbook, schedule: Schedule
    ) -> None:
        """Create per-worker view sheet."""
        ws = wb.create_sheet("By Worker")

        # Headers: Worker, then period columns
        headers = ["Worker"] + [f"P{p.period_index + 1}" for p in schedule.periods]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER

        # Build period assignments lookup
        period_assignments: dict[int, dict[str, list[str]]] = {}
        for period in schedule.periods:
            period_assignments[period.period_index] = {}
            for worker_id, shifts in period.assignments.items():
                period_assignments[period.period_index][worker_id] = [
                    s.shift_type_id for s in shifts
                ]

        # Data rows
        row = 2
        for worker in schedule.workers:
            ws.cell(row=row, column=1, value=f"{worker.name} ({worker.id})")
            for col, period in enumerate(schedule.periods, start=2):
                shift_ids = period_assignments.get(period.period_index, {}).get(
                    worker.id, []
                )
                ws.cell(
                    row=row,
                    column=col,
                    value=", ".join(shift_ids) if shift_ids else "-",
                )
            row += 1

        self._autofit_columns(ws)

    def _autofit_columns(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Auto-fit column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                if column_letter is None:
                    column_letter = get_column_letter(cell.column)
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
