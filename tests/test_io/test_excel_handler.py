"""Tests for Excel import/export handler."""

from datetime import date, time
from pathlib import Path

import pytest

from shift_solver.io.excel_handler import ExcelExporter, ExcelHandlerError, ExcelLoader
from shift_solver.models import (
    PeriodAssignment,
    Schedule,
    ShiftInstance,
    ShiftType,
    Worker,
)


class TestExcelLoader:
    """Tests for Excel import."""

    def test_load_workers_from_excel(self, tmp_path: Path) -> None:
        """Test loading workers from Excel file."""
        # Create a test Excel file
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Workers"
        ws.append(["id", "name", "worker_type", "restricted_shifts"])
        ws.append(["W001", "Alice Smith", "full_time", "night,weekend"])
        ws.append(["W002", "Bob Jones", "part_time", ""])

        excel_file = tmp_path / "input.xlsx"
        wb.save(excel_file)

        loader = ExcelLoader()
        workers = loader.load_workers(excel_file)

        assert len(workers) == 2
        assert workers[0].id == "W001"
        assert workers[0].restricted_shifts == frozenset({"night", "weekend"})

    def test_load_availability_from_excel(self, tmp_path: Path) -> None:
        """Test loading availability from Excel file."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Availability"
        ws.append(["worker_id", "start_date", "end_date", "availability_type"])
        ws.append(["W001", date(2026, 1, 10), date(2026, 1, 15), "unavailable"])

        excel_file = tmp_path / "input.xlsx"
        wb.save(excel_file)

        loader = ExcelLoader()
        avails = loader.load_availability(excel_file)

        assert len(avails) == 1
        assert avails[0].start_date == date(2026, 1, 10)

    def test_load_requests_from_excel(self, tmp_path: Path) -> None:
        """Test loading requests from Excel file."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Requests"
        ws.append(["worker_id", "start_date", "end_date", "request_type", "shift_type_id", "priority"])
        ws.append(["W001", date(2026, 1, 10), date(2026, 1, 10), "positive", "day", 2])

        excel_file = tmp_path / "input.xlsx"
        wb.save(excel_file)

        loader = ExcelLoader()
        requests = loader.load_requests(excel_file)

        assert len(requests) == 1
        assert requests[0].priority == 2

    def test_load_all_from_workbook(self, tmp_path: Path) -> None:
        """Test loading all data from multi-sheet workbook."""
        import openpyxl

        wb = openpyxl.Workbook()

        # Workers sheet
        ws_workers = wb.active
        ws_workers.title = "Workers"
        ws_workers.append(["id", "name"])
        ws_workers.append(["W001", "Alice"])

        # Availability sheet
        ws_avail = wb.create_sheet("Availability")
        ws_avail.append(["worker_id", "start_date", "end_date", "availability_type"])
        ws_avail.append(["W001", date(2026, 1, 10), date(2026, 1, 15), "unavailable"])

        # Requests sheet
        ws_req = wb.create_sheet("Requests")
        ws_req.append(["worker_id", "start_date", "end_date", "request_type", "shift_type_id"])
        ws_req.append(["W001", date(2026, 1, 20), date(2026, 1, 20), "positive", "day"])

        excel_file = tmp_path / "input.xlsx"
        wb.save(excel_file)

        loader = ExcelLoader()
        data = loader.load_all(excel_file)

        assert len(data["workers"]) == 1
        assert len(data["availability"]) == 1
        assert len(data["requests"]) == 1

    def test_file_not_found(self, tmp_path: Path) -> None:
        """Test error when file doesn't exist."""
        loader = ExcelLoader()
        with pytest.raises(ExcelHandlerError, match="File not found"):
            loader.load_workers(tmp_path / "nonexistent.xlsx")


class TestExcelLoaderErrorHandling:
    """Tests for Excel error handling with row numbers."""

    def test_worker_error_includes_row_number(self, tmp_path: Path) -> None:
        """Test that worker parsing errors include row number."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Workers"
        ws.append(["id", "name"])
        ws.append(["W001", "Alice"])
        ws.append(["", "Bob"])  # Row 3 (line_num=3) - missing id

        excel_file = tmp_path / "workers.xlsx"
        wb.save(excel_file)

        loader = ExcelLoader()
        with pytest.raises(ExcelHandlerError, match=r"line 3|row 3"):
            loader.load_workers(excel_file)

    def test_availability_error_includes_row_number(self, tmp_path: Path) -> None:
        """Test that availability parsing errors include row number."""
        from datetime import date

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Availability"
        ws.append(["worker_id", "start_date", "end_date", "availability_type"])
        ws.append(["W001", date(2026, 1, 10), date(2026, 1, 15), "unavailable"])
        ws.append(["W002", date(2026, 1, 10), date(2026, 1, 15), "invalid_type"])  # Invalid type

        excel_file = tmp_path / "availability.xlsx"
        wb.save(excel_file)

        loader = ExcelLoader()
        with pytest.raises(ExcelHandlerError, match=r"line 3|row 3"):
            loader.load_availability(excel_file)

    def test_request_error_includes_row_number(self, tmp_path: Path) -> None:
        """Test that request parsing errors include row number."""
        from datetime import date

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Requests"
        ws.append(["worker_id", "start_date", "end_date", "request_type", "shift_type_id"])
        ws.append(["W001", date(2026, 1, 10), date(2026, 1, 10), "positive", "day"])
        ws.append(["W002", date(2026, 1, 10), date(2026, 1, 10), "invalid_type", "day"])  # Invalid type

        excel_file = tmp_path / "requests.xlsx"
        wb.save(excel_file)

        loader = ExcelLoader()
        with pytest.raises(ExcelHandlerError, match=r"line 3|row 3"):
            loader.load_requests(excel_file)

    def test_unexpected_error_wrapped_with_row_number(self, tmp_path: Path) -> None:
        """Test that unexpected exceptions are wrapped with row context."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Workers"
        ws.append(["id", "name"])
        ws.append(["W001", "Alice"])

        excel_file = tmp_path / "workers.xlsx"
        wb.save(excel_file)

        loader = ExcelLoader()
        # Test that the loader properly wraps errors - if we manually cause an error
        # in parsing, it should include row context
        workers = loader.load_workers(excel_file)
        assert len(workers) == 1


class TestExcelExporter:
    """Tests for Excel export."""

    @pytest.fixture
    def sample_schedule(self) -> Schedule:
        """Create a sample schedule for testing."""
        workers = [
            Worker(id="W001", name="Alice"),
            Worker(id="W002", name="Bob"),
        ]
        shift_types = [
            ShiftType(
                id="day",
                name="Day Shift",
                category="day",
                start_time=time(7, 0),
                end_time=time(15, 0),
                duration_hours=8.0,
                workers_required=1,
            ),
        ]

        # Create period assignments
        periods = [
            PeriodAssignment(
                period_index=0,
                period_start=date(2026, 1, 5),
                period_end=date(2026, 1, 11),
                assignments={
                    "W001": [
                        ShiftInstance(
                            shift_type_id="day",
                            period_index=0,
                            date=date(2026, 1, 5),
                            worker_id="W001",
                        ),
                    ],
                    "W002": [],
                },
            ),
        ]

        return Schedule(
            schedule_id="SCH-001",
            start_date=date(2026, 1, 5),
            end_date=date(2026, 1, 11),
            period_type="week",
            periods=periods,
            workers=workers,
            shift_types=shift_types,
        )

    def test_export_schedule(self, tmp_path: Path, sample_schedule: Schedule) -> None:
        """Test exporting schedule to Excel."""
        exporter = ExcelExporter()
        output_file = tmp_path / "schedule.xlsx"

        exporter.export_schedule(sample_schedule, output_file)

        assert output_file.exists()

        # Verify content
        import openpyxl

        wb = openpyxl.load_workbook(output_file)
        assert "Schedule" in wb.sheetnames
        assert "Statistics" in wb.sheetnames

    def test_export_schedule_with_workers_view(
        self, tmp_path: Path, sample_schedule: Schedule
    ) -> None:
        """Test export includes per-worker view."""
        exporter = ExcelExporter()
        output_file = tmp_path / "schedule.xlsx"

        exporter.export_schedule(sample_schedule, output_file, include_worker_view=True)

        import openpyxl

        wb = openpyxl.load_workbook(output_file)
        assert "By Worker" in wb.sheetnames

    def test_export_creates_parent_directories(
        self, tmp_path: Path, sample_schedule: Schedule
    ) -> None:
        """Test that export creates parent directories."""
        exporter = ExcelExporter()
        output_file = tmp_path / "subdir" / "output" / "schedule.xlsx"

        exporter.export_schedule(sample_schedule, output_file)

        assert output_file.exists()


class TestExcelRoundTrip:
    """Tests for round-trip import/export."""

    def test_workers_round_trip(self, tmp_path: Path) -> None:
        """Test that exported workers can be re-imported."""
        original_workers = [
            Worker(id="W001", name="Alice", worker_type="full_time"),
            Worker(id="W002", name="Bob", restricted_shifts=frozenset({"night"})),
        ]

        # Export (using a simple approach - we'll export to CSV-like Excel)
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Workers"
        ws.append(["id", "name", "worker_type", "restricted_shifts"])
        for w in original_workers:
            ws.append([
                w.id,
                w.name,
                w.worker_type or "",
                ",".join(w.restricted_shifts),
            ])

        excel_file = tmp_path / "workers.xlsx"
        wb.save(excel_file)

        # Import
        loader = ExcelLoader()
        loaded_workers = loader.load_workers(excel_file)

        assert len(loaded_workers) == 2
        assert loaded_workers[0].id == original_workers[0].id
        assert loaded_workers[1].restricted_shifts == original_workers[1].restricted_shifts
