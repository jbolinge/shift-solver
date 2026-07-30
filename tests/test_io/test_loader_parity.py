"""Parity tests between CSVLoader and ExcelLoader.

CSVLoader and ExcelLoader are meant to produce identical domain objects
(Worker, SchedulingRequest, ...) from equivalent input files. Two columns
used to be dropped silently by ExcelLoader even though CSVLoader (mostly)
handled them: worker 'attributes' (both loaders dropped it) and request
'is_hard' (Excel only). These tests pin the fix: both loaders now parse
both columns identically.

Issue: scheduler B3 (loader parity).
"""

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from shift_solver.io.csv_loader import CSVLoader, CSVLoaderError
from shift_solver.io.excel_handler import ExcelHandlerError, ExcelLoader


class TestWorkerAttributesParity:
    """Tests that the 'attributes' worker column is parsed by both loaders."""

    def _create_csv_worker(self, tmp_path: Path, attributes: str) -> Path:
        csv_file = tmp_path / "workers.csv"
        csv_file.write_text(f"id,name,attributes\nworker_1,Worker One,{attributes}\n")
        return csv_file

    def _create_excel_worker(self, tmp_path: Path, attributes) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Workers"
        ws.append(["id", "name", "attributes"])
        ws.append(["worker_1", "Worker One", attributes])
        excel_file = tmp_path / "workers.xlsx"
        wb.save(excel_file)
        return excel_file

    def test_csv_parses_attributes(self, tmp_path: Path) -> None:
        """CSVLoader parses 'key=value;key=value' into Worker.attributes."""
        csv_file = self._create_csv_worker(
            tmp_path, "certification=icu;seniority=senior"
        )
        worker = CSVLoader().load_workers(csv_file)[0]
        assert worker.attributes == {"certification": "icu", "seniority": "senior"}

    def test_excel_parses_attributes(self, tmp_path: Path) -> None:
        """ExcelLoader parses 'key=value;key=value' into Worker.attributes too."""
        excel_file = self._create_excel_worker(
            tmp_path, "certification=icu;seniority=senior"
        )
        worker = ExcelLoader().load_workers(excel_file)[0]
        assert worker.attributes == {"certification": "icu", "seniority": "senior"}

    def test_csv_and_excel_attributes_identical(self, tmp_path: Path) -> None:
        """Equivalent CSV and Excel worker rows produce the same attributes."""
        csv_file = self._create_csv_worker(tmp_path, "certification=icu")
        excel_file = self._create_excel_worker(tmp_path, "certification=icu")

        csv_worker = CSVLoader().load_workers(csv_file)[0]
        excel_worker = ExcelLoader().load_workers(excel_file)[0]

        assert (
            csv_worker.attributes == excel_worker.attributes == {"certification": "icu"}
        )

    def test_missing_attributes_column_defaults_empty_both(
        self, tmp_path: Path
    ) -> None:
        """No 'attributes' column at all yields an empty dict for both loaders."""
        csv_file = tmp_path / "workers.csv"
        csv_file.write_text("id,name\nworker_1,Worker One\n")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Workers"
        ws.append(["id", "name"])
        ws.append(["worker_1", "Worker One"])
        excel_file = tmp_path / "workers.xlsx"
        wb.save(excel_file)

        csv_worker = CSVLoader().load_workers(csv_file)[0]
        excel_worker = ExcelLoader().load_workers(excel_file)[0]

        assert csv_worker.attributes == excel_worker.attributes == {}

    def test_malformed_attributes_rejected_by_both(self, tmp_path: Path) -> None:
        """An entry missing '=' is rejected by both loaders."""
        csv_file = self._create_csv_worker(tmp_path, "not-a-pair")
        excel_file = self._create_excel_worker(tmp_path, "not-a-pair")

        with pytest.raises(CSVLoaderError, match="Invalid attributes entry"):
            CSVLoader().load_workers(csv_file)

        with pytest.raises(ExcelHandlerError, match="Invalid attributes entry"):
            ExcelLoader().load_workers(excel_file)


class TestRequestIsHardParity:
    """Tests that the 'is_hard' request column is parsed by both loaders."""

    def _create_csv_request(self, tmp_path: Path, is_hard: str) -> Path:
        csv_file = tmp_path / "requests.csv"
        csv_file.write_text(
            "worker_id,start_date,end_date,request_type,shift_type_id,is_hard\n"
            f"worker_1,2026-01-10,2026-01-10,positive,day,{is_hard}\n"
        )
        return csv_file

    def _create_excel_request(self, tmp_path: Path, is_hard) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Requests"
        ws.append(
            [
                "worker_id",
                "start_date",
                "end_date",
                "request_type",
                "shift_type_id",
                "is_hard",
            ]
        )
        ws.append(
            [
                "worker_1",
                date(2026, 1, 10),
                date(2026, 1, 10),
                "positive",
                "day",
                is_hard,
            ]
        )
        excel_file = tmp_path / "requests.xlsx"
        wb.save(excel_file)
        return excel_file

    def test_excel_parses_is_hard_true(self, tmp_path: Path) -> None:
        """ExcelLoader no longer silently drops 'is_hard' (was always None)."""
        excel_file = self._create_excel_request(tmp_path, "true")
        request = ExcelLoader().load_requests(excel_file)[0]
        assert request.is_hard is True

    def test_excel_parses_is_hard_false(self, tmp_path: Path) -> None:
        excel_file = self._create_excel_request(tmp_path, "no")
        request = ExcelLoader().load_requests(excel_file)[0]
        assert request.is_hard is False

    def test_excel_missing_is_hard_is_none(self, tmp_path: Path) -> None:
        excel_file = self._create_excel_request(tmp_path, None)
        request = ExcelLoader().load_requests(excel_file)[0]
        assert request.is_hard is None

    def test_csv_and_excel_is_hard_identical(self, tmp_path: Path) -> None:
        """Equivalent CSV and Excel request rows produce the same is_hard."""
        csv_file = self._create_csv_request(tmp_path, "true")
        excel_file = self._create_excel_request(tmp_path, "true")

        csv_request = CSVLoader().load_requests(csv_file)[0]
        excel_request = ExcelLoader().load_requests(excel_file)[0]

        assert csv_request.is_hard is excel_request.is_hard is True

    def test_invalid_is_hard_rejected_by_both(self, tmp_path: Path) -> None:
        csv_file = self._create_csv_request(tmp_path, "maybe")
        excel_file = self._create_excel_request(tmp_path, "maybe")

        with pytest.raises(CSVLoaderError, match="Invalid is_hard value"):
            CSVLoader().load_requests(csv_file)

        with pytest.raises(ExcelHandlerError, match="Invalid is_hard value"):
            ExcelLoader().load_requests(excel_file)


class TestDateFormatConstructorParity:
    """Tests that both loaders honor a constructor-level date_format."""

    def test_csv_loader_honors_explicit_eu_format(self, tmp_path: Path) -> None:
        """An ambiguous date is interpreted per the loader's date_format."""
        csv_file = tmp_path / "availability.csv"
        csv_file.write_text(
            "worker_id,start_date,end_date,availability_type\n"
            "worker_1,01/02/2026,01/03/2026,unavailable\n"
        )
        avail = CSVLoader(date_format="eu").load_availability(csv_file)[0]
        assert avail.start_date == date(2026, 2, 1)  # EU: day/month
        assert avail.end_date == date(2026, 3, 1)

    def test_excel_loader_honors_explicit_eu_format(self, tmp_path: Path) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Availability"
        ws.append(["worker_id", "start_date", "end_date", "availability_type"])
        ws.append(["worker_1", "01/02/2026", "01/03/2026", "unavailable"])
        excel_file = tmp_path / "availability.xlsx"
        wb.save(excel_file)

        avail = ExcelLoader(date_format="eu").load_availability(excel_file)[0]
        assert avail.start_date == date(2026, 2, 1)
        assert avail.end_date == date(2026, 3, 1)

    def test_default_date_format_is_auto_for_both(self, tmp_path: Path) -> None:
        """Constructing without date_format keeps the historical "auto" behavior."""
        csv_file = tmp_path / "availability.csv"
        csv_file.write_text(
            "worker_id,start_date,end_date,availability_type\n"
            "worker_1,2026-01-15,2026-01-20,unavailable\n"
        )
        avail = CSVLoader().load_availability(csv_file)[0]
        assert avail.start_date == date(2026, 1, 15)
