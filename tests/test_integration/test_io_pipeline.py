"""Integration tests for the I/O pipeline: CSV/Excel import -> solve -> export."""

from datetime import date
from pathlib import Path

import pytest

from factories import create_period_dates
from shift_solver.io import (
    CSVLoader,
    ExcelExporter,
    ExcelLoader,
    SampleGenerator,
)
from shift_solver.models import ShiftType, Worker
from shift_solver.solver import ShiftSolver
from shift_solver.validation import ScheduleValidator


@pytest.mark.integration
class TestCSVImportSolveExport:
    """Test complete CSV import -> solve -> export workflow."""

    def test_csv_import_solve_export_workflow(self, tmp_path: Path) -> None:
        """Test full workflow: generate CSV -> import -> solve -> export."""
        # Step 1: Generate sample CSV data
        data_dir = tmp_path / "data"
        generator = SampleGenerator(industry="retail", seed=42)
        start_date = date(2026, 2, 2)
        end_date = date(2026, 2, 15)  # 2 weeks
        generator.generate_to_csv(
            output_dir=data_dir,
            num_workers=10,
            start_date=start_date,
            end_date=end_date,
        )

        # Step 2: Import from CSV
        loader = CSVLoader()
        workers = loader.load_workers(data_dir / "workers.csv")
        availability = loader.load_availability(data_dir / "availability.csv")
        requests = loader.load_requests(data_dir / "requests.csv")

        assert len(workers) == 10
        assert all(isinstance(w, Worker) for w in workers)

        # Step 3: Create shift types from generator
        shift_types = generator.generate_shift_types()
        period_dates = create_period_dates(
            start_date=start_date,
            num_periods=2,
        )

        # Step 4: Solve
        solver = ShiftSolver(
            workers=workers,
            shift_types=shift_types,
            period_dates=period_dates,
            schedule_id="CSV-PIPELINE-TEST",
            availabilities=availability,
            requests=requests,
        )

        result = solver.solve(time_limit_seconds=60)
        assert result.success
        assert result.schedule is not None

        # Step 5: Validate
        validator = ScheduleValidator(
            schedule=result.schedule,
            availabilities=availability,
            requests=requests,
        )
        validation = validator.validate()
        assert validation.is_valid

        # Step 6: Export to Excel
        output_file = tmp_path / "schedule.xlsx"
        exporter = ExcelExporter()
        exporter.export_schedule(result.schedule, output_file)

        assert output_file.exists()
        assert output_file.stat().st_size > 0

    def test_csv_with_restrictions_respected(self, tmp_path: Path) -> None:
        """Test that worker restrictions from CSV are respected in solution."""
        # Create CSV with specific restrictions
        workers_csv = tmp_path / "workers.csv"
        workers_csv.write_text(
            "id,name,worker_type,restricted_shifts\n"
            "W001,Alice,full_time,night\n"
            "W002,Bob,full_time,\n"
            "W003,Charlie,full_time,night\n"
        )

        loader = CSVLoader()
        workers = loader.load_workers(workers_csv)

        from datetime import time

        shift_types = [
            ShiftType(
                id="day",
                name="Day Shift",
                category="day",
                start_time=time(9, 0),
                end_time=time(17, 0),
                duration_hours=8.0,
                workers_required=1,
            ),
            ShiftType(
                id="night",
                name="Night Shift",
                category="night",
                start_time=time(22, 0),
                end_time=time(6, 0),
                duration_hours=8.0,
                workers_required=1,
                is_undesirable=True,
            ),
        ]

        period_dates = create_period_dates(num_periods=2)

        solver = ShiftSolver(
            workers=workers,
            shift_types=shift_types,
            period_dates=period_dates,
            schedule_id="RESTRICTION-TEST",
        )

        result = solver.solve(time_limit_seconds=30)
        assert result.success
        assert result.schedule is not None

        # Verify W001 and W003 are never on night shift
        for period in result.schedule.periods:
            for worker_id, shifts in period.assignments.items():
                if worker_id in ["W001", "W003"]:
                    for shift in shifts:
                        assert shift.shift_type_id != "night", (
                            f"{worker_id} assigned to restricted night shift"
                        )


@pytest.mark.integration
class TestExcelRoundtrip:
    """Test Excel import/export roundtrip data integrity."""

    def test_excel_data_integrity_roundtrip(self, tmp_path: Path) -> None:
        """Test that data survives Excel export -> re-import."""
        # Step 1: Generate sample data
        generator = SampleGenerator(industry="healthcare", seed=123)
        start_date = date(2026, 2, 2)

        workers = generator.generate_workers(8)
        shift_types = generator.generate_shift_types()
        period_dates = create_period_dates(start_date=start_date, num_periods=3)

        # Step 2: Solve
        solver = ShiftSolver(
            workers=workers,
            shift_types=shift_types,
            period_dates=period_dates,
            schedule_id="ROUNDTRIP-TEST",
        )

        result = solver.solve(time_limit_seconds=60)
        assert result.success
        schedule = result.schedule
        assert schedule is not None

        # Step 3: Export to Excel
        export_file = tmp_path / "schedule.xlsx"
        exporter = ExcelExporter()
        exporter.export_schedule(schedule, export_file)

        # Step 4: Verify file structure
        import openpyxl

        wb = openpyxl.load_workbook(export_file)
        assert "Schedule" in wb.sheetnames
        assert "Statistics" in wb.sheetnames
        assert "By Worker" in wb.sheetnames

        # Step 5: Verify data in Statistics sheet
        stats_sheet = wb["Statistics"]
        rows = list(stats_sheet.iter_rows(values_only=True))
        assert rows[0][0] == "Worker"  # Header check

        # Count total assignments from exported data
        exported_totals = {}
        for row in rows[1:]:
            if row[0]:
                worker_id = row[0]
                total = row[-1]  # Last column is Total
                exported_totals[worker_id] = total

        # Compare with original schedule
        original_totals = {}
        for worker in workers:
            original_totals[worker.id] = 0
            for period in schedule.periods:
                original_totals[worker.id] += len(
                    period.get_worker_shifts(worker.id)
                )

        for worker_id in original_totals:
            assert exported_totals.get(worker_id) == original_totals[worker_id], (
                f"Mismatch for {worker_id}: "
                f"exported={exported_totals.get(worker_id)}, "
                f"original={original_totals[worker_id]}"
            )

    def test_excel_import_all_sheets(self, tmp_path: Path) -> None:
        """Test loading all data from a multi-sheet Excel workbook."""
        # Generate Excel file with sample data
        generator = SampleGenerator(industry="warehouse", seed=456)
        excel_file = tmp_path / "sample_data.xlsx"
        generator.generate_to_excel(
            output_file=excel_file,
            num_workers=12,
            start_date=date(2026, 2, 2),
            end_date=date(2026, 2, 28),
        )

        # Load all sheets
        loader = ExcelLoader()
        data = loader.load_all(excel_file)

        assert "workers" in data
        assert "availability" in data
        assert "requests" in data

        assert len(data["workers"]) == 12
        assert all(isinstance(w, Worker) for w in data["workers"])


@pytest.mark.integration
class TestMixedFormats:
    """Test mixing CSV and Excel in the same workflow."""

    def test_csv_workers_excel_export(self, tmp_path: Path) -> None:
        """Test importing CSV workers and exporting to Excel."""
        # Create workers CSV
        workers_csv = tmp_path / "workers.csv"
        workers_csv.write_text(
            "id,name,worker_type\n"
            "W001,Alice Smith,full_time\n"
            "W002,Bob Jones,full_time\n"
            "W003,Charlie Brown,part_time\n"
            "W004,Diana Ross,full_time\n"
        )

        loader = CSVLoader()
        workers = loader.load_workers(workers_csv)

        from datetime import time

        shift_types = [
            ShiftType(
                id="day",
                name="Day Shift",
                category="day",
                start_time=time(9, 0),
                end_time=time(17, 0),
                duration_hours=8.0,
                workers_required=1,
            ),
        ]

        period_dates = create_period_dates(num_periods=2)

        solver = ShiftSolver(
            workers=workers,
            shift_types=shift_types,
            period_dates=period_dates,
            schedule_id="MIXED-FORMAT",
        )

        result = solver.solve(time_limit_seconds=30)
        assert result.success

        # Export to Excel
        output = tmp_path / "output.xlsx"
        exporter = ExcelExporter()
        exporter.export_schedule(result.schedule, output)

        assert output.exists()


@pytest.mark.integration
@pytest.mark.smoke
class TestIOSmoke:
    """Quick smoke tests for I/O operations."""

    def test_csv_loader_smoke(
        self,
        sample_workers_csv: Path,
        sample_availability_csv: Path,
        sample_requests_csv: Path,
    ) -> None:
        """Smoke test for CSV loading."""
        loader = CSVLoader()

        workers = loader.load_workers(sample_workers_csv)
        assert len(workers) > 0

        availability = loader.load_availability(sample_availability_csv)
        assert isinstance(availability, list)

        requests = loader.load_requests(sample_requests_csv)
        assert isinstance(requests, list)

    def test_sample_generator_smoke(self, tmp_path: Path) -> None:
        """Smoke test for sample generation."""
        generator = SampleGenerator(industry="retail", seed=1)

        workers = generator.generate_workers(5)
        assert len(workers) == 5

        shift_types = generator.generate_shift_types()
        assert len(shift_types) > 0

        # Generate to CSV
        generator.generate_to_csv(
            output_dir=tmp_path,
            num_workers=5,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 1, 14),
        )

        assert (tmp_path / "workers.csv").exists()
        assert (tmp_path / "shift_types.csv").exists()
        assert (tmp_path / "availability.csv").exists()
        assert (tmp_path / "requests.csv").exists()
