"""Tests for I/O CLI commands."""

from datetime import date
from pathlib import Path

import pytest
from click.testing import CliRunner

from shift_solver.cli.main import cli


class TestGenerateSamplesCommand:
    """Tests for generate-samples command."""

    def test_generate_samples_csv(self, tmp_path: Path) -> None:
        """Test generating CSV sample files."""
        runner = CliRunner()
        result = runner.invoke(
            cli,
            [
                "generate-samples",
                "--output-dir",
                str(tmp_path),
                "--industry",
                "retail",
                "--num-workers",
                "5",
                "--months",
                "1",
                "--format",
                "csv",
            ],
        )

        assert result.exit_code == 0, result.output
        assert "Generating retail sample data" in result.output
        assert (tmp_path / "workers.csv").exists()
        assert (tmp_path / "shift_types.csv").exists()

    def test_generate_samples_excel(self, tmp_path: Path) -> None:
        """Test generating Excel sample file."""
        runner = CliRunner()
        result = runner.invoke(
            cli,
            [
                "generate-samples",
                "--output-dir",
                str(tmp_path),
                "--industry",
                "healthcare",
                "--num-workers",
                "5",
                "--months",
                "1",
                "--format",
                "excel",
            ],
        )

        assert result.exit_code == 0, result.output
        assert (tmp_path / "sample_data.xlsx").exists()

    def test_generate_samples_both_formats(self, tmp_path: Path) -> None:
        """Test generating both CSV and Excel."""
        runner = CliRunner()
        result = runner.invoke(
            cli,
            [
                "generate-samples",
                "--output-dir",
                str(tmp_path),
                "--format",
                "both",
                "--num-workers",
                "3",
                "--months",
                "1",
            ],
        )

        assert result.exit_code == 0, result.output
        assert (tmp_path / "csv" / "workers.csv").exists()
        assert (tmp_path / "excel" / "sample_data.xlsx").exists()

    def test_generate_samples_with_seed(self, tmp_path: Path) -> None:
        """Test that seed produces reproducible output."""
        runner = CliRunner()

        # Generate with seed
        result1 = runner.invoke(
            cli,
            [
                "generate-samples",
                "--output-dir",
                str(tmp_path / "run1"),
                "--seed",
                "42",
                "--num-workers",
                "3",
                "--months",
                "1",
            ],
        )
        assert result1.exit_code == 0

        result2 = runner.invoke(
            cli,
            [
                "generate-samples",
                "--output-dir",
                str(tmp_path / "run2"),
                "--seed",
                "42",
                "--num-workers",
                "3",
                "--months",
                "1",
            ],
        )
        assert result2.exit_code == 0

        # Files should be identical
        workers1 = (tmp_path / "run1" / "workers.csv").read_text()
        workers2 = (tmp_path / "run2" / "workers.csv").read_text()
        assert workers1 == workers2


class TestImportDataCommand:
    """Tests for import-data command."""

    def test_import_workers_csv(self, tmp_path: Path) -> None:
        """Test importing workers from CSV."""
        # Create a test CSV file
        workers_csv = tmp_path / "workers.csv"
        workers_csv.write_text("id,name,worker_type\nW001,Alice,full_time\nW002,Bob,part_time\n")

        runner = CliRunner()
        result = runner.invoke(
            cli,
            ["import-data", "--workers", str(workers_csv)],
        )

        assert result.exit_code == 0, result.output
        assert "Loaded 2 workers" in result.output

    def test_import_availability_csv(self, tmp_path: Path) -> None:
        """Test importing availability from CSV."""
        avail_csv = tmp_path / "availability.csv"
        avail_csv.write_text(
            "worker_id,start_date,end_date,availability_type\n"
            "W001,2026-01-10,2026-01-15,unavailable\n"
        )

        runner = CliRunner()
        result = runner.invoke(
            cli,
            ["import-data", "--availability", str(avail_csv)],
        )

        assert result.exit_code == 0, result.output
        assert "Loaded 1 availability records" in result.output

    def test_import_excel_workbook(self, tmp_path: Path) -> None:
        """Test importing from Excel workbook."""
        import openpyxl

        wb = openpyxl.Workbook()

        # Workers sheet
        ws = wb.active
        ws.title = "Workers"
        ws.append(["id", "name"])
        ws.append(["W001", "Alice"])
        ws.append(["W002", "Bob"])

        # Availability sheet
        ws_avail = wb.create_sheet("Availability")
        ws_avail.append(["worker_id", "start_date", "end_date", "availability_type"])
        ws_avail.append(["W001", date(2026, 1, 10), date(2026, 1, 15), "unavailable"])

        excel_file = tmp_path / "data.xlsx"
        wb.save(excel_file)

        runner = CliRunner()
        result = runner.invoke(
            cli,
            ["import-data", "--excel", str(excel_file)],
        )

        assert result.exit_code == 0, result.output
        assert "Workers: 2" in result.output
        assert "Availability records: 1" in result.output

    def test_import_no_files_error(self) -> None:
        """Test error when no files specified."""
        runner = CliRunner()
        result = runner.invoke(cli, ["import-data"])

        assert result.exit_code != 0
        assert "No input files specified" in result.output

    def test_import_invalid_csv_error(self, tmp_path: Path) -> None:
        """Test error on invalid CSV."""
        bad_csv = tmp_path / "bad.csv"
        bad_csv.write_text("wrong,columns\nfoo,bar\n")

        runner = CliRunner()
        result = runner.invoke(
            cli,
            ["import-data", "--workers", str(bad_csv)],
        )

        assert result.exit_code != 0
        assert "Worker import error" in result.output


class TestExportCommand:
    """Tests for export command."""

    @pytest.fixture
    def sample_schedule_json(self, tmp_path: Path) -> Path:
        """Create a sample schedule JSON file."""
        import json

        schedule_data = {
            "schedule_id": "SCH-TEST",
            "start_date": "2026-01-05",
            "end_date": "2026-01-11",
            "periods": [
                {
                    "period_index": 0,
                    "period_start": "2026-01-05",
                    "period_end": "2026-01-11",
                    "assignments": {
                        "W001": [
                            {"shift_type_id": "day", "date": "2026-01-05"},
                            {"shift_type_id": "day", "date": "2026-01-06"},
                        ],
                        "W002": [
                            {"shift_type_id": "night", "date": "2026-01-05"},
                        ],
                    },
                }
            ],
            "statistics": {},
        }

        schedule_file = tmp_path / "schedule.json"
        with open(schedule_file, "w") as f:
            json.dump(schedule_data, f)

        return schedule_file

    def test_export_to_excel(
        self, tmp_path: Path, sample_schedule_json: Path
    ) -> None:
        """Test exporting schedule to Excel."""
        output_file = tmp_path / "output.xlsx"

        runner = CliRunner()
        result = runner.invoke(
            cli,
            [
                "export",
                "--schedule",
                str(sample_schedule_json),
                "--output",
                str(output_file),
                "--format",
                "excel",
            ],
        )

        assert result.exit_code == 0, result.output
        assert output_file.exists()
        assert "Schedule exported to" in result.output

    def test_export_to_json(
        self, tmp_path: Path, sample_schedule_json: Path
    ) -> None:
        """Test exporting schedule to JSON."""
        output_file = tmp_path / "output.json"

        runner = CliRunner()
        result = runner.invoke(
            cli,
            [
                "export",
                "--schedule",
                str(sample_schedule_json),
                "--output",
                str(output_file),
                "--format",
                "json",
            ],
        )

        assert result.exit_code == 0, result.output
        assert output_file.exists()

        # Verify JSON content
        import json

        with open(output_file) as f:
            data = json.load(f)
        assert data["schedule_id"] == "SCH-TEST"

    def test_export_excel_no_worker_view(
        self, tmp_path: Path, sample_schedule_json: Path
    ) -> None:
        """Test exporting Excel without worker view."""
        output_file = tmp_path / "output.xlsx"

        runner = CliRunner()
        result = runner.invoke(
            cli,
            [
                "export",
                "--schedule",
                str(sample_schedule_json),
                "--output",
                str(output_file),
                "--no-worker-view",
            ],
        )

        assert result.exit_code == 0, result.output

        # Verify no "By Worker" sheet
        import openpyxl

        wb = openpyxl.load_workbook(output_file)
        assert "By Worker" not in wb.sheetnames
